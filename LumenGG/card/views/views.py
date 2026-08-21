from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, Http404, JsonResponse
from django.contrib import messages
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Case, When, IntegerField, Avg, BooleanField, Min, Count
from django.db.models.functions import Cast
from django.conf import settings
from django.core import signing
from django.utils import timezone
from django.views.decorators.http import require_POST

from ..models import Card, Character, Tag, CardComment
from battlelog.game.catalog import _card_snapshot
from battlelog.game.effects import EffectResolutionError
from battlelog.game.engine import EngineError
from battlelog.game.sandbox import (
    EVENT_LABELS as SANDBOX_EVENT_LABELS,
    ZONE_LABELS as SANDBOX_ZONE_LABELS,
    EffectSandboxError,
    continue_effect_sandbox,
    describe_ability_choices,
    project_effect_sandbox,
    sandbox_event_options,
    sandbox_prototype_abilities,
    sandbox_prototype_definition,
    start_effect_sandbox,
)
from battlelog.game.schema import validate_effect_definition
from battlelog.game.spec import ALL_ZONES, PHASES, TRIGGERS
from collection.models import CollectionCard, Pack
from deck.models import Deck
from ..forms import CardForm, TagCreateForm, CardTagEditForm, CardCreateForm, CardUpdateForm, CardTranslationUpdateForm, CardCommentForm
from ..effect_review import CardEffectReviewForm, card_effect_review_context
from ..search import card_matches_search, card_matches_search_exact
from decorators import permission_required
import re, random, os, json, uuid
from io import BytesIO
from zipfile import BadZipFile
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from PIL import Image
from common.language import (
    DEFAULT_LANGUAGE,
    LANGUAGE_ENGLISH,
    LANGUAGE_JAPANESE,
    SUPPORTED_LANGUAGES,
    get_language,
    normalize_language,
    render_localized_markup,
    translated_character_field,
    ui_text,
)

NEUTRAL_CHARACTER_ID = 1
DETAIL_TEXT_IMPORT_DIR = 'card_detail_text_imports'
DETAIL_TEXT_IMPORT_REQUIRED_COLUMNS = ['첫 출전팩', '번호', '이름', '보충 설명']
EFFECT_SANDBOX_SIGNING_SALT = 'card.effect-sandbox.v1'
EFFECT_SANDBOX_MAX_AGE_SECONDS = 60 * 60
EFFECT_SANDBOX_MAX_REQUEST_BYTES = 512 * 1024


def _detail_text_import_path(token):
    if not re.fullmatch(r'[0-9a-f]{32}', token or ''):
        raise ValueError('잘못된 업로드 토큰입니다.')
    return f'{DETAIL_TEXT_IMPORT_DIR}/{token}.xlsx'


def _normalize_excel_value(value):
    if value is None:
        return ''
    if isinstance(value, str):
        text = value
    else:
        text = str(value)
    return text.replace('\r\n', '\n').replace('\r', '\n').strip()


def _normalize_card_number(value):
    if isinstance(value, int):
        return str(value).zfill(3)
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).zfill(3)

    text = _normalize_excel_value(value)
    if text.isdigit():
        return text.zfill(3)
    return text.upper()


def _build_detail_text_import_code(pack, number):
    pack_code = _normalize_excel_value(pack).upper()
    card_number = _normalize_card_number(number)
    if not pack_code or not card_number:
        return ''
    return f'{pack_code}-{card_number}'


def _save_detail_text_import_file(uploaded_file):
    if not uploaded_file:
        raise ValueError('xlsx 파일을 선택해 주세요.')
    if not uploaded_file.name.lower().endswith('.xlsx'):
        raise ValueError('xlsx 파일만 업로드할 수 있습니다.')

    token = uuid.uuid4().hex
    path = _detail_text_import_path(token)
    default_storage.save(path, uploaded_file)
    return token


def _delete_detail_text_import_file(token):
    try:
        path = _detail_text_import_path(token)
    except ValueError:
        return
    if default_storage.exists(path):
        default_storage.delete(path)


def _load_detail_text_import_workbook(token):
    path = _detail_text_import_path(token)
    if not default_storage.exists(path):
        raise ValueError('업로드된 파일을 찾을 수 없습니다. 다시 업로드해 주세요.')
    with default_storage.open(path, 'rb') as uploaded:
        return openpyxl.load_workbook(BytesIO(uploaded.read()), read_only=True, data_only=True)


def _get_detail_text_import_sheet_names(token):
    workbook = _load_detail_text_import_workbook(token)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def _get_detail_text_header_indexes(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError('헤더 행을 찾을 수 없습니다.')

    headers = {}
    for index, value in enumerate(header_row):
        header = _normalize_excel_value(value)
        if header:
            headers[header] = index

    missing_columns = [column for column in DETAIL_TEXT_IMPORT_REQUIRED_COLUMNS if column not in headers]
    if missing_columns:
        raise ValueError('필수 컬럼이 없습니다: ' + ', '.join(missing_columns))
    return headers


def _get_detail_text_import_cell(row, headers, column):
    index = headers.get(column)
    if index is None or index >= len(row):
        return ''
    return _normalize_excel_value(row[index])


def _load_detail_text_import_card_maps():
    cards = Card.objects.only('id', 'code', 'name', 'detail_text')
    cards_by_code = {card.code: card for card in cards if card.code}
    cards_by_name = {}
    duplicate_names = set()

    for card in cards:
        if not card.name:
            continue
        if card.name in cards_by_name:
            duplicate_names.add(card.name)
        else:
            cards_by_name[card.name] = card

    for name in duplicate_names:
        cards_by_name.pop(name, None)

    return cards_by_code, cards_by_name, duplicate_names


def parse_detail_text_import(token, sheet_name):
    workbook = _load_detail_text_import_workbook(token)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError('선택한 시트를 찾을 수 없습니다.')

        sheet = workbook[sheet_name]
        headers = _get_detail_text_header_indexes(sheet)
        cards_by_code, cards_by_name, duplicate_names = _load_detail_text_import_card_maps()
        rows = []
        summary = {
            'total_rows': 0,
            'matched_count': 0,
            'update_count': 0,
            'no_change_count': 0,
            'name_fallback_count': 0,
            'unmatched_count': 0,
            'empty_detail_count': 0,
            'warning_count': 0,
        }

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            pack = _get_detail_text_import_cell(row, headers, '첫 출전팩')
            number = row[headers['번호']] if headers['번호'] < len(row) else None
            excel_number = _normalize_excel_value(number)
            excel_name = _get_detail_text_import_cell(row, headers, '이름')
            detail_text = _get_detail_text_import_cell(row, headers, '보충 설명')
            card_type = _get_detail_text_import_cell(row, headers, '종류')
            effect_text = _get_detail_text_import_cell(row, headers, '효과')

            if not any([pack, excel_number, excel_name, detail_text, card_type, effect_text]):
                continue

            summary['total_rows'] += 1
            code = _build_detail_text_import_code(pack, number)
            card = cards_by_code.get(code)
            match_method = 'code' if card else ''
            warning = ''

            if not card and excel_name:
                if excel_name in duplicate_names:
                    warning = '같은 이름의 카드가 여러 장 있어 이름으로 매칭하지 않았습니다.'
                else:
                    card = cards_by_name.get(excel_name)
                    if card:
                        match_method = 'name'
                        summary['name_fallback_count'] += 1

            if not detail_text:
                status = 'empty_detail'
                status_label = '빈 보충 설명'
                status_class = 'muted'
                summary['empty_detail_count'] += 1
                changed = False
            elif not card:
                status = 'unmatched'
                status_label = '미매칭'
                status_class = 'danger'
                summary['unmatched_count'] += 1
                changed = False
            else:
                summary['matched_count'] += 1
                if match_method == 'code' and excel_name and card.name != excel_name:
                    warning = f'코드는 매칭되었지만 DB 이름은 "{card.name}"입니다.'
                    summary['warning_count'] += 1

                changed = (card.detail_text or '') != detail_text
                if changed:
                    status = 'will_update'
                    status_label = '변경 예정'
                    status_class = 'accent'
                    summary['update_count'] += 1
                else:
                    status = 'no_change'
                    status_label = '변경 없음'
                    status_class = 'success'
                    summary['no_change_count'] += 1

                if match_method == 'name':
                    status_label = '이름 매칭 / ' + status_label
                    status_class = 'warning' if changed else 'success'

            if warning and not (card and match_method == 'code' and excel_name and card.name != excel_name):
                summary['warning_count'] += 1

            rows.append({
                'row_number': row_number,
                'code': code,
                'excel_name': excel_name,
                'db_name': card.name if card else '',
                'card_id': card.id if card else None,
                'match_method': match_method,
                'status': status,
                'status_label': status_label,
                'status_class': status_class,
                'changed': changed,
                'detail_text': detail_text,
                'type': card_type,
                'effect_text': effect_text,
                'warning': warning,
            })

        return {
            'sheet_name': sheet_name,
            'rows': rows,
            'summary': summary,
        }
    finally:
        workbook.close()


def apply_detail_text_import(token, sheet_name):
    preview = parse_detail_text_import(token, sheet_name)
    changed_rows = [row for row in preview['rows'] if row['changed'] and row['card_id']]
    cards_by_id = Card.objects.in_bulk([row['card_id'] for row in changed_rows])
    cards_to_update = []

    with transaction.atomic():
        for row in changed_rows:
            card = cards_by_id.get(row['card_id'])
            if not card:
                continue
            card.detail_text = row['detail_text']
            cards_to_update.append(card)
        if cards_to_update:
            Card.objects.bulk_update(cards_to_update, ['detail_text'])

    return preview, len(cards_to_update)

# Create your views here.
def index(req, template_name='card/list.html'):
    language = get_language(req)
    page = req.GET.get('page', '1')
    
    form = CardForm(req.GET, language=language)
    
    if not form.is_valid():
        char = None
        stype = None
        pos = None
        body = None
        specialpos = None
        specialtype = None
        keyword = None
        pack = None
        framenum = None
        frametype = None
        sort = None
        ultimate = None
    else:
        char = form.cleaned_data['char']
        stype = form.cleaned_data['type']
        pos = form.cleaned_data['pos']
        body = form.cleaned_data['body']
        specialpos = form.cleaned_data['specialpos']
        specialtype = form.cleaned_data['specialtype']
        keyword = form.cleaned_data['keyword']
        pack = form.cleaned_data['pack']
        framenum = form.cleaned_data['framenum']
        frametype = form.cleaned_data['frametype']
        sort = form.cleaned_data['sort']
        ultimate = form.cleaned_data['ultimate']
    
    q = Q()
    
    if char: 
        q.add(Q(character__in=char), q.AND)
    if stype: q.add(Q(type__in=stype), q.AND)
    if ultimate: q.add(Q(ultimate=True), q.AND)
    if pos: q.add(Q(pos__in=pos), q.AND)
    if body: 
        if body == "없음":
            q.add(Q(body__isnull=True) | Q(body=""), q.AND)
        else: q.add(Q(body__in=body), q.AND)
    if specialtype: 
        qtemp = Q()
        for sp in specialtype:
            qtemp.add(Q(special__contains=sp), q.OR)
        q.add(qtemp, q.AND)
    if specialpos: 
        qtemp = Q()
        for sp in specialpos:
            qtemp.add(Q(special__contains=sp), qtemp.OR)
        q.add(qtemp, q.AND)
    if pack: q.add(Q(code__contains=pack), q.AND)
    if framenum:
        if frametype == '일치': q.add(Q(frame=int(framenum)), q.AND)
        elif frametype == '이상': q.add(Q(frame__gte=int(framenum)), q.AND)
        elif frametype == '이하': q.add(Q(frame__lte=int(framenum)), q.AND)
    data = Card.objects.filter(q).select_related('character').prefetch_related(
        'translations',
        'character__translations',
    ).distinct().annotate(avgscore=Avg('comments__score'))

    if keyword:
        matched_ids = [
            card.id for card in data
            if card_matches_search(card, keyword, include_keywords=True)
        ]
        data = data.filter(id__in=matched_ids)
    
    if sort:
        if '히트' in sort:
            data = data.filter(Q(type='공격') & ~Q(hit='콤보'))
        elif '카운터' in sort:
            data = data.filter(Q(type='공격') & ~Q(counter='콤보'))
        elif '가드' in sort:
            data = data.filter(Q(type='공격') & ~Q(guard='X'))
        
        if sort == '-속도':
            data = data.order_by('-frame')
        elif sort == '+속도':
            data = data.order_by('frame')
        elif sort == '-데미지':
            data = data.order_by('-damage')
        elif sort == '+데미지':
            data = data.order_by('damage')
        elif sort == "-평점":
            data = data.order_by('-avgscore')
        elif sort == "+평점":
            data = data.order_by('avgscore')
        elif sort == '-히트':
            data = data.order_by(Cast('hit', IntegerField()).desc())
        elif sort == '+히트':
            data = data.order_by(Cast('hit', IntegerField()))
        elif sort == '-카운터':
            data = data.order_by(Cast('counter', IntegerField()).desc())
        elif sort == '+카운터':
            data = data.order_by(Cast('counter', IntegerField()))
        elif sort == '-가드':
            data = data.order_by(Cast('guard', IntegerField()).desc())
        elif sort == '+가드':
            data = data.order_by(Cast('guard', IntegerField()))
        elif sort == '출시일':
            data = data.annotate(earliest_release=Min('collection_card__pack__released')).order_by('-earliest_release')
        elif sort == '+출시일':
            data = data.annotate(earliest_release=Min('collection_card__pack__released')).order_by('earliest_release')
    else:
        data = data.order_by('id')
    
    if req.GET.get('random', None):
        data = random.choices(k=int(req.GET.get('random', 1)), population=data, replace=False)
    
    codes = Pack.objects.filter(released__gt=timezone.now())
    for code in codes:
        data = data.annotate(
            unReleased = Case(When(code__contains=code.code, then=True), default=False)
        )
    
    paginator = Paginator(data, 12)
    page_data = paginator.get_page(page)
    
    form = CardForm(req.GET, language=language)
    
    
    context = {
        'form': form,
        'cards': page_data
    }
    return render(req, template_name, context=context)

def indexV2(req):
    return index(req, 'card/list_v2.html')

def get_card_adoption_stats(card, language='ko'):
    deck_scope = Deck.objects.filter(
        deleted=False,
    ).annotate(
        card_entry_count=Count('cids', distinct=True),
    ).filter(
        card_entry_count__gte=15,
    )

    is_neutral_card = card.character_id == NEUTRAL_CHARACTER_ID
    if is_neutral_card:
        scope_label = ui_text('전체 덱', language)
    else:
        deck_scope = deck_scope.filter(character_id=card.character_id)
        character_name = translated_character_field(card.character, language, 'name')
        language = normalize_language(language)
        if language == LANGUAGE_ENGLISH:
            scope_label = f'{character_name} decks'
        elif language == LANGUAGE_JAPANESE:
            scope_label = f'{character_name}デッキ'
        else:
            scope_label = f'{character_name} 덱'

    total_decks = deck_scope.count()
    adopted_decks = deck_scope.filter(cids__card_id=card.id).distinct().count()
    adoption_rate = adopted_decks / total_decks * 100 if total_decks else 0

    return {
        'is_neutral_card': is_neutral_card,
        'scope_label': scope_label,
        'total_decks': total_decks,
        'adopted_decks': adopted_decks,
        'adoption_rate': adoption_rate,
    }


def card_tag_edit_context(card, language):
    language = normalize_language(language)
    source_values = {
        'keyword': card.keyword or '',
        'search': card.search or '',
        'hidden': card.hiddenKeyword or '',
    }
    if language == DEFAULT_LANGUAGE:
        return {
            'values': source_values,
            'placeholders': {'keyword': '', 'search': '', 'hidden': ''},
            'is_translation': False,
        }

    translation = card.translations.filter(language=language).first()
    return {
        'values': {
            'keyword': (translation.keyword if translation else '') or '',
            'search': (translation.search if translation else '') or '',
            'hidden': (translation.hiddenKeyword if translation else '') or '',
        },
        'placeholders': source_values,
        'is_translation': True,
    }

def detail(req, id=0, template_name='card/detail.html'):
    language = get_language(req)
    try:
        card = Card.objects.select_related('character').prefetch_related(
            'translations',
            'character__translations',
        ).get(id = id)
    except Card.DoesNotExist:
        raise Http404("카드가 존재하지 않습니다.")
    
    relation = {}
    
    kws = card.search.split('/')[:-1]
    for kw in kws:
        relation[kw] = Card.objects.select_related('character').prefetch_related(
            'translations',
            'character__translations',
        ).filter(keyword__contains=kw)
        relation[kw] = relation[kw].exclude(id = id)
    
    cc = CollectionCard.objects.select_related('pack').prefetch_related('pack__translations').filter(card = card)
    cc = cc.annotate(
        custom_order=Case(
        When(rare='N', then=0),
        When(rare='SR', then=1),
        When(rare='EXR', then=2),
        When(rare='AN', then=3),
        When(rare='AEX', then=4),
        When(rare='SAR', then=5),
        When(rare='SP', then=6),
        default=7,
        output_field=IntegerField(),
        )
    )
    cc = cc.order_by('pack__released', 'code', 'custom_order')
    
    if cc.exists() and cc[0].pack.released > timezone.now().date():
        unReleased = True
    else:
        unReleased = False
    
    context = {
        'card': card,
        'relation': relation,
        'cc': cc,
        'unReleased': unReleased,
        'adoption_stats': get_card_adoption_stats(card, language),
        'tag_edit': card_tag_edit_context(card, language),
    }
    return render(req, template_name, context=context)

def detailV2(req, id=0):
    return detail(req, id, 'card/detail_v2.html')


def _next_effect_review_card(current_pk):
    queryset = Card.objects.exclude(
        effect_definition__reviewed=True,
    ).exclude(pk=current_pk).order_by('pk')
    return queryset.filter(pk__gt=current_pk).first() or queryset.first()


@permission_required('card.change_card')
def effectReview(req, id=0):
    card = get_object_or_404(
        Card.objects.select_related('character').prefetch_related('qna'),
        pk=id,
    )
    if req.method == 'POST':
        form = CardEffectReviewForm(req.POST, instance=card)
        if form.is_valid():
            card = form.save()
            messages.success(
                req,
                f'{card.code or ""} {card.name}의 자동 효과 정의를 저장했습니다.',
            )
            if '_saveandnextunreviewed' in req.POST:
                next_card = _next_effect_review_card(card.pk)
                if next_card:
                    return redirect('card:effectReview', next_card.pk)
                messages.success(req, '남은 미검수 카드가 없습니다.')
            return redirect('card:effectReview', card.pk)
    else:
        form = CardEffectReviewForm(instance=card)

    next_card = _next_effect_review_card(card.pk)
    definition = card.effect_definition if isinstance(card.effect_definition, dict) else {}
    sandbox_abilities = []
    ability_groups = (
        (sandbox_prototype_abilities(), True),
        (definition.get('abilities') or [], False),
    )
    for abilities, is_prototype in ability_groups:
        for ability in abilities:
            choice_description = describe_ability_choices(ability)
            sandbox_abilities.append({
                'id': str(ability.get('id') or ''),
                'label': str(
                    ability.get('label') or ability.get('draft_text')
                    or ability.get('id') or '이름 없는 효과'
                ),
                'mode': str(ability.get('mode') or ''),
                'events': sandbox_event_options(ability),
                'active_zones': list(ability.get('active_zones') or []),
                'choice_steps': choice_description['steps'],
                'automatic_steps': choice_description['automatic_steps'],
                'choice_warnings': choice_description['warnings'],
                'prototype': is_prototype,
            })
    context = {
        'card': card,
        'form': form,
        'review': card_effect_review_context(card),
        'sandbox_abilities': sandbox_abilities,
        'sandbox_cards': list(
            Card.objects.exclude(pk=card.pk).exclude(code__isnull=True).order_by(
                'code', 'pk',
            ).values('id', 'code', 'name', 'type', 'frame', 'damage', 'pos')
        ),
        'sandbox_events': [
            {'value': event, 'label': SANDBOX_EVENT_LABELS.get(event, event)}
            for event in sorted(TRIGGERS)
        ],
        'sandbox_zones': [
            {'value': zone, 'label': SANDBOX_ZONE_LABELS.get(zone, zone)}
            for zone in ALL_ZONES
        ],
        'sandbox_phases': PHASES,
        'next_unreviewed_card': next_card,
        'unreviewed_card_count': Card.objects.exclude(
            effect_definition__reviewed=True,
        ).count(),
    }
    return render(req, 'card/effect_review_v2.html', context)


def _effect_sandbox_json(req):
    if len(req.body) > EFFECT_SANDBOX_MAX_REQUEST_BYTES:
        raise EffectSandboxError('효과 테스트 요청이 너무 큽니다.')
    try:
        value = json.loads(req.body.decode('utf-8') or '{}')
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise EffectSandboxError('효과 테스트 요청 JSON이 올바르지 않습니다.') from exc
    if not isinstance(value, dict):
        raise EffectSandboxError('효과 테스트 요청은 객체여야 합니다.')
    return value


def _effect_sandbox_response(payload):
    token_payload = {
        **payload,
    }
    token = signing.dumps(
        token_payload,
        salt=EFFECT_SANDBOX_SIGNING_SALT,
        compress=True,
    )
    return JsonResponse({
        'ok': True,
        'token': token,
        'result': project_effect_sandbox(payload),
    })


def _effect_sandbox_error(exc, *, status=400):
    return JsonResponse({'ok': False, 'error': str(exc)}, status=status)


@permission_required('card.change_card')
@require_POST
def effectSandboxStart(req, id=0):
    card = get_object_or_404(Card.objects.select_related('character'), pk=id)
    try:
        body = _effect_sandbox_json(req)
        ability_id = body.get('ability_id')
        definition = (
            sandbox_prototype_definition(ability_id)
            or body.get('effect_definition', card.effect_definition)
        )
        issues = validate_effect_definition(
            definition,
            card_has_text=bool((card.text or '').strip()),
        )
        if issues:
            first = issues[0]
            raise EffectSandboxError(
                f'효과 정의를 먼저 수정해 주세요: {first.path} {first.message}'
            )
        config = body.get('config')
        if not isinstance(config, dict):
            raise EffectSandboxError('테스트 상황 설정이 필요합니다.')
        placements = config.get('cards') if isinstance(config.get('cards'), list) else []
        support_ids = {
            int(item.get('card_id'))
            for item in placements
            if isinstance(item, dict) and str(item.get('card_id') or '').isdigit()
        }
        if card.pk in support_ids:
            raise EffectSandboxError('검수 중인 카드는 추가 카드로 다시 배치할 수 없습니다.')
        support_cards = list(
            Card.objects.select_related('character').filter(pk__in=support_ids)
        )
        if len(support_cards) != len(support_ids):
            raise EffectSandboxError('배치하려는 카드 중 현재 DB에 없는 카드가 있습니다.')
        support_snapshots = {
            str(item.pk): _card_snapshot(item) for item in support_cards
        }
        payload = start_effect_sandbox(
            _card_snapshot(card), definition, ability_id,
            support_snapshots, config,
        )
        payload['card_id'] = card.pk
        payload['reviewer_id'] = req.user.pk
        return _effect_sandbox_response(payload)
    except (EffectSandboxError, EffectResolutionError, EngineError) as exc:
        return _effect_sandbox_error(exc)


@permission_required('card.change_card')
@require_POST
def effectSandboxDecision(req, id=0):
    get_object_or_404(Card, pk=id)
    try:
        body = _effect_sandbox_json(req)
        try:
            payload = signing.loads(
                str(body.get('token') or ''),
                salt=EFFECT_SANDBOX_SIGNING_SALT,
                max_age=EFFECT_SANDBOX_MAX_AGE_SECONDS,
            )
        except signing.SignatureExpired as exc:
            raise EffectSandboxError('효과 테스트가 만료되었습니다. 다시 실행해 주세요.') from exc
        except signing.BadSignature as exc:
            raise EffectSandboxError('효과 테스트 상태 서명이 올바르지 않습니다.') from exc
        if payload.get('card_id') != id or payload.get('reviewer_id') != req.user.pk:
            raise EffectSandboxError('현재 카드와 사용자의 효과 테스트가 아닙니다.')
        selected = body.get('selected')
        if not isinstance(selected, list):
            raise EffectSandboxError('선택 결과는 배열이어야 합니다.')
        payload = continue_effect_sandbox(payload, selected)
        return _effect_sandbox_response(payload)
    except (EffectSandboxError, EffectResolutionError, EngineError) as exc:
        return _effect_sandbox_error(exc)

def detailName(req, name):
    language = get_language(req)
    try:
        card = Card.objects.get(name=name)
    except Card.DoesNotExist:
        candidates = Card.objects.select_related('character').prefetch_related(
            'translations',
            'character__translations',
        ).order_by('id')
        for candidate in candidates:
            if card_matches_search_exact(candidate, name, include_keywords=False):
                return detailV2(req, candidate.id)

        cards = [
            candidate for candidate in candidates
            if card_matches_search(candidate, name, include_keywords=True)
        ]
        context = {
            'cards': cards
        }
        return render(req, 'card/card404.html', context=context)
    else:
        return detailV2(req, card.id)

def detailNameLegacy(req, name):
    try:
        card = Card.objects.get(name=name)
    except Card.DoesNotExist:
        candidates = Card.objects.select_related('character').prefetch_related(
            'translations',
            'character__translations',
        ).order_by('id')
        for candidate in candidates:
            if card_matches_search_exact(candidate, name, include_keywords=False):
                return detail(req, candidate.id)
        cards = [
            candidate for candidate in candidates
            if card_matches_search(candidate, name, include_keywords=True)
        ]
        context = {
            'cards': cards
        }
        return render(req, 'card/card404.html', context=context)
    else:
        return detail(req, card.id)

@permission_required('card.add_card')
def create(req, template_name='card/create.html', detail_route='card:detail'):
    if req.method == 'GET':
        form = CardCreateForm()
        
        return render(req, template_name, context={'form': form, 'is_update': False})
    else:
        form = CardCreateForm(req.POST, req.FILES)
        if form.is_valid():
            try:
                card = Card.objects.get(name = form.cleaned_data['name'])
                return redirect(detail_route, card.id)
            except:
                card = form.save(commit=False)
                uploaded_image = req.FILES.get('imageFile')
                if uploaded_image:
                    save_card_image_files(card, uploaded_image)
                card.save()
                
            for r in form.data.getlist('rare'):
                newCC = CollectionCard(
                    card = card,
                    pack = Pack.objects.get(id = form.data.get('pack')),
                    code = card.code,
                    image = card.img,
                    img_sm = card.img_mid,
                    name = card.name,
                    rare = r)
                #print(newCC.__dict__)
                newCC.save()
            return redirect(detail_route, card.id)
        else:
            return render(req, template_name, context={'form': form, 'is_update': False})

def createV2(req):
    return create(req, 'card/create_v2.html', 'card:detail')


def _card_translation_preview(form, language):
    return {
        field_name: render_localized_markup(form[field_name].value() or '', language)
        for field_name in ('text', 'detail_text')
        if field_name in form.fields
    }


@permission_required('card.change_card')
def update(req, id=0, template_name='card/create.html', detail_route='card:detail'):
    language = get_language(req)
    language = normalize_language(language)
    is_translation_update = language != DEFAULT_LANGUAGE
    language_labels = dict(SUPPORTED_LANGUAGES)

    try:
        card = Card.objects.get(id=id)
    except Card.DoesNotExist:
        raise Http404("카드가 존재하지 않습니다.")

    context_base = {
        'card': card,
        'is_update': True,
        'is_translation_update': is_translation_update,
        'edit_language': language,
        'edit_language_label': language_labels.get(language, language),
    }

    if is_translation_update:
        translation = card.translations.filter(language=language).first()

        if req.method == 'GET':
            form = CardTranslationUpdateForm(instance=translation, card=card, language=language)
            return render(req, template_name, context={
                **context_base,
                'form': form,
                'translation_preview': _card_translation_preview(form, language),
            })

        form = CardTranslationUpdateForm(req.POST, instance=translation, card=card, language=language)
        if form.is_valid():
            form.save()
            return redirect(detail_route, card.id)
        return render(req, template_name, context={
            **context_base,
            'form': form,
            'translation_preview': _card_translation_preview(form, language),
        })
    
    if req.method == 'GET':
        form = CardUpdateForm(instance=card)
        return render(req, template_name, context={**context_base, 'form': form})
    
    form = CardUpdateForm(req.POST, req.FILES, instance=card)
    if form.is_valid():
        card = form.save(commit=False)
        uploaded_image = req.FILES.get('imageFile')
        if uploaded_image:
            save_card_image_files(card, uploaded_image)
        card.save()
        CollectionCard.objects.filter(card=card).update(name=card.name)
        return redirect(detail_route, card.id)
    return render(req, template_name, context={**context_base, 'form': form})

def updateV2(req, id=0):
    return update(req, id, 'card/create_v2.html', 'card:detail')


@permission_required('card.change_card')
def detailTextImport(req):
    context = {
        'sheet_names': [],
        'selected_sheet': '',
        'token': '',
        'preview': None,
    }

    if req.method == 'POST':
        action = req.POST.get('action', '')

        try:
            if action == 'upload':
                token = _save_detail_text_import_file(req.FILES.get('xlsx_file'))
                try:
                    sheet_names = _get_detail_text_import_sheet_names(token)
                except Exception:
                    _delete_detail_text_import_file(token)
                    raise

                context.update({
                    'token': token,
                    'sheet_names': sheet_names,
                    'selected_sheet': sheet_names[0] if sheet_names else '',
                })
                messages.success(req, '파일을 업로드했습니다. 가져올 시트를 선택해 주세요.')

            elif action == 'preview':
                token = req.POST.get('token', '')
                sheet_name = req.POST.get('sheet_name', '')
                preview = parse_detail_text_import(token, sheet_name)
                context.update({
                    'token': token,
                    'sheet_names': _get_detail_text_import_sheet_names(token),
                    'selected_sheet': sheet_name,
                    'preview': preview,
                })

            elif action == 'apply':
                token = req.POST.get('token', '')
                sheet_name = req.POST.get('sheet_name', '')
                preview, updated_count = apply_detail_text_import(token, sheet_name)
                _delete_detail_text_import_file(token)
                messages.success(req, f'{updated_count}장의 보충 설명을 저장했습니다.')
                context.update({
                    'preview': preview,
                    'selected_sheet': sheet_name,
                })

            else:
                messages.error(req, '알 수 없는 요청입니다.')

        except ValueError as error:
            messages.error(req, str(error))
        except (InvalidFileException, BadZipFile):
            messages.error(req, '올바른 xlsx 파일을 업로드해 주세요.')

    return render(req, 'card/detail_text_import_v2.html', context=context)


def save_card_image_files(card, uploaded_image):
    path = os.path.join(settings.MEDIA_ROOT, 'webp', (card.code+'.webp'))
    handle_uploaded_file(uploaded_image, path)
    
    compress_image(path, os.path.join(settings.MEDIA_ROOT, 
                                    'webpsm', 
                                    (card.code+'.webp')), 319, 100)
    compress_image(path, os.path.join(settings.MEDIA_ROOT, 
                                    'webpmin', 
                                    (card.code+'.webp')), 213, 100)
    card.img = 'https://images.hinoto.kr/lumendb/webp/' + card.code + '.webp'
    card.img_mid = 'https://images.hinoto.kr/lumendb/webpsm/' + card.code + '.webp'
    card.img_sm = 'https://images.hinoto.kr/lumendb/webpmin/' + card.code + '.webp'

def handle_uploaded_file(f, filePath):
    print(filePath)
    os.makedirs(os.path.dirname(filePath), exist_ok=True)
    with open(filePath, "wb+") as destination:
        for chunk in f.chunks():
            destination.write(chunk)

#mid = 235, #sm = 157
def compress_image(filePath: str, newFilePath: str, max_width: int = 235, quality = 100):
    with Image.open(filePath) as img:
        width, height = img.size
        if width > max_width:
            img.thumbnail((max_width, max_width))
        img.save(newFilePath, quality=quality)
            
def tagList(req):
    page = req.GET.get('page', '1')
    keyword = req.GET.get('keyword', '')
    
    q = Q()
    q.add(Q(name__contains=keyword), q.OR)
    q.add(Q(description__contains=keyword), q.OR)
    tags = Tag.objects.filter(q)
    
    paginator = Paginator(tags, 30)
    page_data = paginator.get_page(page)
    
    if req.method == 'GET':
        return render(req, 'card/tagList.html', context={'tags': page_data})

def tagDetail(req, id=0):
    try:
        tag = Tag.objects.get(id = id)
    except Tag.DoesNotExist:
        raise Http404("태그가 존재하지 않습니다.")
    
    keyword = Card.objects.filter(keyword__contains=tag.name)
    search = Card.objects.filter(search__contains=tag.name)
    
    context = {
        'tag': tag,
        'keyword': keyword,
        'search': search,
    }
    return render(req, 'card/tagDetail.html', context=context)

@permission_required('card.tag_update')
def tagCreate(req):
    if req.method == 'POST':
        form = TagCreateForm(req.POST)
        if form.is_valid():
            tag = form.save()
            return redirect('card:tagDetail', tag.id)
    else:
        form = TagCreateForm()
    
    return render(req, 'card/tagCreate.html', context={'form': form})

@permission_required('card.tag_update')
def tagUpdate(req, id=0):
    try:
        tag = Tag.objects.get(id = id)
    except Tag.DoesNotExist:
        raise Http404("태그가 존재하지 않습니다.")
    
    if req.method == 'POST':
        form = TagCreateForm(req.POST, instance=tag)
        if form.is_valid():
            tag = form.save()
            return render(req, 'card/tagDetail.html', context={'tag': tag})
    else:
        form = TagCreateForm(instance=tag)
    
    return render(req, 'card/tagUpdate.html', context={'form': form})


@permission_required('card.tag_update')
def editCardTag(req, id=0):
    language = get_language(req)
    language = normalize_language(language)
    try:
        card = Card.objects.get(id = id)
    except Card.DoesNotExist:
        raise Http404("카드가 존재하지 않습니다.")

    if language == DEFAULT_LANGUAGE:
        card.hiddenKeyword = req.POST['hidden']
        card.keyword = req.POST['keyword']
        card.search = req.POST['search']
        card.save(update_fields=['hiddenKeyword', 'keyword', 'search'])
    else:
        translation = card.translations.filter(language=language).first()
        if translation is None:
            translation = card.translations.create(language=language)
        translation.hiddenKeyword = req.POST['hidden']
        translation.keyword = req.POST['keyword']
        translation.search = req.POST['search']
        translation.save(update_fields=['hiddenKeyword', 'keyword', 'search'])
    
    return redirect('card:detail', id)

def comment(req, id=0):
    try:
        card = Card.objects.get(id = id)
    except Card.DoesNotExist:
        raise Http404("카드가 존재하지 않습니다.")
    
    comments = CardComment.objects.filter(card=card).order_by('-created_at')
    
    if req.method == 'POST':
        if not req.user.is_authenticated:
            return redirect('card:comment', card.id)
        data = json.loads(req.body)
        try:
            com = comments.get(author=req.user)
            com.score = data['score']
            com.comment = data['comment']
        except CardComment.DoesNotExist:
            com = CardComment(
                author = req.user,
                score = data['score'],
                comment = data['comment'],
                card = card
            )
        com.save()
        return JsonResponse({'status': 100 })
    else:
        try:
            com = comments.get(author=req.user)
        except:
            form = CardCommentForm()
        else:
            form = CardCommentForm(instance=com)
        
    page = req.GET.get('page', '1')
    
    context = {
        'card': card,
        'comments': comments,
        'form': form,
    }
    return render(req, 'card/comment.html', context = context)

def commentList(req):
    page = req.GET.get('page', '1')
    
    comments = CardComment.objects.order_by('-created_at')
    paginator = Paginator(comments, 30)
    page_data = paginator.get_page(page)
    
    context = {
        'comments': page_data
    }
    return render(req, 'card/commentList.html', context=context)
